from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


DEFAULT_COLUMNS = ["task_id", "name", "keyword", "status", "remark"]
STATUS_PENDING = "\u5f85\u5904\u7406"


@dataclass
class TaskTable:
    rows: list[dict[str, Any]]
    columns: list[str]
    source_path: Path
    key_column: str = "name"
    status_column: str = "status"
    remark_column: str = "remark"

    def ensure_columns(self) -> None:
        for column in DEFAULT_COLUMNS:
            if column not in self.columns:
                self.columns.append(column)
        for row in self.rows:
            for column in self.columns:
                row.setdefault(column, "")
            if not row.get("keyword"):
                row["keyword"] = row.get(self.key_column, "")
            if not row.get(self.status_column):
                row[self.status_column] = STATUS_PENDING

    def selectable_rows(self, only_statuses: set[str] | None, skip_statuses: set[str] | None) -> Iterable[dict[str, Any]]:
        for row in self.rows:
            status = str(row.get(self.status_column, "")).strip()
            if only_statuses and status not in only_statuses:
                continue
            if skip_statuses and status in skip_statuses:
                continue
            yield row

    def update(self, row: dict[str, Any], status: str | None = None, remark: str | None = None) -> None:
        if status is not None:
            row[self.status_column] = status
        if remark is not None:
            row[self.remark_column] = remark

    def save_xlsx(self, path: str | Path) -> Path:
        from openpyxl import Workbook

        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "result"
        sheet.append(self.columns)
        for row in self.rows:
            sheet.append([row.get(column, "") for column in self.columns])
        workbook.save(path)
        return path

    def save_csv(self, path: str | Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=self.columns)
            writer.writeheader()
            writer.writerows(self.rows)
        return path


def _read_csv(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row) for row in reader]
        columns = list(reader.fieldnames or DEFAULT_COLUMNS)
    return rows, columns


def _read_txt(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    rows = []
    for index, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        name = line.strip()
        if not name:
            continue
        rows.append({"task_id": f"{index:03d}", "name": name, "keyword": name, "status": STATUS_PENDING, "remark": ""})
    return rows, list(DEFAULT_COLUMNS)


def _read_xlsx(path: Path, sheet_name: str | None = None) -> tuple[list[dict[str, Any]], list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return [], list(DEFAULT_COLUMNS)
    columns = [str(value).strip() if value is not None else "" for value in values[0]]
    rows = []
    for values_row in values[1:]:
        if not any(value is not None and str(value).strip() for value in values_row):
            continue
        rows.append({columns[i]: values_row[i] if i < len(values_row) else "" for i in range(len(columns))})
    return rows, columns


def load_task_table(data_source: dict[str, Any]) -> TaskTable:
    path = Path(data_source["file"])
    if not path.exists():
        raise FileNotFoundError(f"Data source not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows, columns = _read_csv(path)
    elif suffix == ".txt":
        rows, columns = _read_txt(path)
    elif suffix == ".xlsx":
        rows, columns = _read_xlsx(path, data_source.get("sheet"))
    else:
        raise ValueError(f"Unsupported data source type: {path.suffix}")

    table = TaskTable(
        rows=rows,
        columns=columns,
        source_path=path,
        key_column=data_source.get("key_column", "name"),
        status_column=data_source.get("status_column", "status"),
        remark_column=data_source.get("remark_column", "remark"),
    )
    table.ensure_columns()
    return table
